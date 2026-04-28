"""Export the briefing research table to XLSX."""

from __future__ import annotations

from pathlib import Path

from src.schemas import ArticleSummary

XLSX_COLUMNS: tuple[str, ...] = (
    "PDF Source",
    "주제",
    "내용 요약",
    "기사 제목",
    "기사 원본URL",
    "기사 출처",
    "기사 내용 정리",
    "결론 및 시사점",
)


class XLSXExportError(RuntimeError):
    """Raised when XLSX output cannot be created."""


def export_briefing_xlsx(rows: tuple[ArticleSummary, ...], output_path: Path) -> Path:
    """Write the final issue/article table as an XLSX file."""
    if not rows:
        raise XLSXExportError("at least one row is required")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise XLSXExportError("openpyxl package is not installed") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "briefing"
    worksheet.append(XLSX_COLUMNS)

    for row in rows:
        worksheet.append(
            [
                row.pdf_source,
                row.topic,
                row.pdf_summary,
                row.title,
                row.url,
                row.source,
                row.summary,
                row.conclusion,
            ]
        )

    header_fill = PatternFill("solid", fgColor="7F7F7F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = (32, 24, 48, 42, 48, 20, 78, 56)
    for index, width in enumerate(widths, start=1):
        column = worksheet.cell(row=1, column=index).column_letter
        worksheet.column_dimensions[column].width = width

    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 120
    worksheet.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path
